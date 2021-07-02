# -*- coding: UTF-8 -*-
# ! /usr/bin/python3

import openpyxl


class Excel(object):
    def __init__(self, filepath):
        # 1.打开 Excel 表格并获取表格名称
        self.workbook = openpyxl.load_workbook(filepath)
        self.sheet_names = self.workbook.sheetnames
        # 设置工作表缓存
        self._sheet_cache = {}

    def get_rows(self, sheet_name):
        """
        获取工作表中行数据列表
        :param sheet_name:
        :return:
        """
        table = self.get_sheet(sheet_name)
        return table.rows

    def get_cell(self, sheet_name, row, column):
        """
        获取工作表单元格数据
        :param sheet_name: 工作表名称
        :param row: {int} row
        :param column: {int} column
        :return:
        """
        sheet = self.get_sheet(sheet_name)
        return sheet.cell(row=row, column=column).value

    def get_sheet(self, sheet_name):
        """
        获取工作表数据
        :param sheet_name: sheet_name
        :return: sheet 数据
        """
        if sheet_name in self._sheet_cache:
            return self._sheet_cache[sheet_name]
        else:
            self._sheet_cache[sheet_name] = self.workbook[sheet_name]
            return self._sheet_cache[sheet_name]

    def get_thead(self, sheet_name):
        f"""
        获取工作表表头
        :param sheet_name: 工作表名称 
        :return: {dict[str, int]}
        """
        sheet = self.get_sheet(sheet_name)
        column_total = sheet.max_column

        # 获取列名
        thead = {}
        for i in range(1, column_total + 1):
            name = sheet.cell(row=1, column=i).value
            thead[name] = i

        return thead

    def close(self):
        self.workbook.close()
